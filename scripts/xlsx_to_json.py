#!/usr/bin/env python3
"""Convert an .xlsx workbook to JSON rows for the Node migration script."""
from __future__ import annotations

import datetime as dt
import json
import sys

from openpyxl import load_workbook


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.date().isoformat() if isinstance(value, dt.datetime) else value.isoformat()
    return value


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: xlsx_to_json.py /path/to/workbook.xlsx", file=sys.stderr)
        return 2

    workbook = load_workbook(sys.argv[1], read_only=True, data_only=True)
    result = {}
    for sheet in workbook.worksheets:
        rows = []
        max_col = sheet.max_column or 0
        for row in sheet.iter_rows(min_row=1, max_col=max_col, values_only=True):
            rows.append([normalize_cell(value) for value in row])
        result[sheet.title] = rows

    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
