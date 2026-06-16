from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


OUTPUT = "出货检验报告.xlsx"


thin = Side(style="thin", color="000000")
medium = Side(style="medium", color="000000")
no_side = Side(style=None)


def border(all_side=thin):
    return Border(left=all_side, right=all_side, top=all_side, bottom=all_side)


def mixed_border(left=thin, right=thin, top=thin, bottom=thin):
    return Border(left=left, right=right, top=top, bottom=bottom)


def set_cell(ws, cell, value, font_size=9, bold=False, align="center", valign="center", wrap=True):
    c = ws[cell]
    c.value = value
    c.font = Font(name="宋体", size=font_size, bold=bold, color="000000")
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap, shrink_to_fit=False)
    return c


def merge_set(ws, ref, value, font_size=9, bold=False, align="center", valign="center", wrap=True):
    ws.merge_cells(ref)
    cell = ref.split(":")[0]
    return set_cell(ws, cell, value, font_size, bold, align, valign, wrap)


def apply_border(ws, ref, side=thin):
    cells = ws[ref]
    for row in cells:
        for c in row:
            c.border = border(side)


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "出货检验报告"
    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.35, right=0.35, top=0.28, bottom=0.32, header=0.15, footer=0.12)
    ws.print_options.horizontalCentered = True
    ws.print_area = "A1:K31"

    widths = {
        "A": 10.0,
        "B": 14.0,
        "C": 12.5,
        "D": 7.0,
        "E": 7.2,
        "F": 6.2,
        "G": 7.0,
        "H": 16.0,
        "I": 6.2,
        "J": 10.5,
        "K": 5.6,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for r in range(1, 39):
        ws.row_dimensions[r].height = 13.2
    ws.row_dimensions[1].height = 9
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 15
    ws.row_dimensions[6].height = 12.5
    ws.row_dimensions[7].height = 12.5
    ws.row_dimensions[8].height = 15
    ws.row_dimensions[9].height = 15
    ws.row_dimensions[36].height = 18
    ws.row_dimensions[37].height = 20
    ws.row_dimensions[38].height = 20

    merge_set(ws, "A2:K2", "广东风华高新科技股份有限公司", 16, True)
    ws["A2"].font = Font(name="黑体", size=16, bold=True)
    merge_set(ws, "A3:K3", "出货检验报告", 10, False)
    ws["A3"].font = Font(name="宋体", size=10, bold=False)
    merge_set(ws, "A4:K4", "FENGHUA  ADVANCED  TECHNOLOGY  HOLDING  CO., LTD.\nOUTGOING  INSPECTION  SHEET", 9.5, False)
    ws["A4"].font = Font(name="Times New Roman", size=9.5, bold=False)

    merge_set(ws, "A5:C5", "客户名称 (Customer): __________________________", 7, False, "left")
    merge_set(ws, "D5:E5", "日期 (Date): 2025-12-24", 7, False, "left")
    merge_set(ws, "F5:H5", "温/湿度 (IR, T/R.H.): 22℃ / 65%", 7, False, "left")
    merge_set(ws, "I5:K5", "可焊性条件 (Solder ability): 245±5℃  2±0.5S", 7, False, "left")

    headers = [
        ("A6:A7", "客户型号\n(Customer P/N)"),
        ("B6:B7", "型号规格\n(Part No.)"),
        ("C6:C7", "批号\nLot No."),
        ("D6:D6", "绝缘电阻\nIR"),
        ("E6:E6", "容量\nCap"),
        ("F6:F6", "损耗DF/Q值\nD.F.\n(x10⁻⁴)"),
        ("G6:G6", "耐电压\nDWV"),
        ("H6:J6", "检验条件 (Condition)"),
        ("K6:K7", "结论\nResult"),
    ]
    for ref, text in headers:
        c = merge_set(ws, ref, text, 6.0, False)
    subheaders = {
        "D7": "标准值\nSpec.",
        "E7": "范围\nRange",
        "F7": "标准值\nSpec.",
        "G7": "标准值\nSpec.",
        "H7": "容量 / 损耗角正切 / Q值\nCap/DF/Q",
        "I7": "绝缘电阻IR",
        "J7": "耐电压DWV",
    }
    for cell, text in subheaders.items():
        set_cell(ws, cell, text, 6.0, False)

    data = [
        [
            ("A8:A8", "GJM1555C1HR80WB01DF"),
            ("B8:B8", "GJM1555C1HR80WB01D"),
            ("C8:C8", ""),
            ("D8:D8", "IR≥10000MΩ"),
            ("E8:E8", "0.75~0.85pF"),
            ("F8:F8", "Q≥416"),
            ("G8:G8", "150V\n1~5s"),
            ("H8:H8", "1.0±0.1MHz /\n0.5~5.0Vrms"),
            ("I8:I8", "50V"),
            ("J8:J8", "150V, 50mA\nMAX"),
            ("K8:K8", "合格"),
        ],
        [
            ("A9:A9", ""),
            ("B9:B9", ""),
            ("C9:C9", ""),
            ("D9:D9", ""),
            ("E9:E9", ""),
            ("F9:F9", ""),
            ("G9:G9", ""),
            ("H9:H9", ""),
            ("I9:I9", ""),
            ("J9:J9", ""),
            ("K9:K9", ""),
        ],
    ]
    for row in data:
        for ref, text in row:
            merge_set(ws, ref, text, 6.2)

    apply_border(ws, "A6:K9", thin)

    signoffs = [
        ("B20", "检验员\nOperated By", "B21", "梁起煜"),
        ("D20", "审核\nApproved\nBy", "D21", "合 格"),
        ("F20", "日期\nDate", "F21", "2025-12-24"),
        ("H20:I20", "报告编号\nRep No", "H21:I21", "PBW202512241311"),
        ("J20:K20", "页码", "J21:K21", "第  1  页    共  1  页"),
    ]
    for label_ref, label, value_ref, value in signoffs:
        if ":" in label_ref:
            merge_set(ws, label_ref, label, 7)
            merge_set(ws, value_ref, value, 7)
        else:
            set_cell(ws, label_ref, label, 7)
            set_cell(ws, value_ref, value, 7)
    ws["D21"].font = Font(name="宋体", size=8.5, bold=True, color="C00000")

    for row in ws.iter_rows(min_row=1, max_row=39, min_col=1, max_col=11):
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "center",
                    vertical="center",
                    wrap_text=True,
                    shrink_to_fit=False,
                )

    ws.freeze_panes = None

    basis = wb.create_sheet("合格依据")
    basis.sheet_view.showGridLines = False
    basis.column_dimensions["A"].width = 18
    basis.column_dimensions["B"].width = 24
    basis.column_dimensions["C"].width = 42
    basis.column_dimensions["D"].width = 58
    basis.column_dimensions["E"].width = 22
    for r in range(1, 14):
        basis.row_dimensions[r].height = 30

    title = basis["A1"]
    title.value = "GJM1555C1HR80WB01D/DF 合格参数依据"
    title.font = Font(name="黑体", size=14, bold=True)
    basis.merge_cells("A1:E1")
    title.alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        ["项目", "出货报告填写值", "计算/判定", "依据", "建议检验条件"],
        ["型号识别", "GJM1555C1HR80WB01DF", "渠道查询匹配官方基础料号 GJM1555C1HR80WB01D；DF 可能为客户/包装后缀。", "Murata/DigiKey 均以 GJM1555C1HR80WB01D 展示基础料号。", "按 GJM1555C1HR80WB01D 控制"],
        ["容量 Cap.", "0.75~0.85pF", "0.8pF ±0.05pF，因此下限 0.75pF，上限 0.85pF。", "Murata rated value: 0.8pF ±0.05pF。", "1.0±0.1MHz, 0.5~5.0Vrms"],
        ["额定电压 R.V.", "50V", "额定直流电压为 50V。", "Murata rated value / DigiKey product attribute: 50V。", "不得超过额定电压"],
        ["介质/温度特性", "C0G/NP0", "C0G: 0±30ppm/℃，工作温度 -55~125℃。", "Murata rated value: C0G(EIA), 0±30ppm/℃, -55~125℃。", "按 C0G 规格判定"],
        ["Q / DF", "Q≥416", "Murata 规定 Q≥400+20C，C=0.8pF，故 Q≥416；等效 DF≤1/416≈0.0024。", "Murata specifications and test methods: Q≥400+20C。", "1.0±0.1MHz, 0.5~5.0Vrms"],
        ["绝缘电阻 I.R.", "IR≥10000MΩ", "室温下两端测量，绝缘电阻大于 10000MΩ。", "Murata specifications: More than 10000MΩ; measurement voltage rated voltage; charging time 1min。", "50V, 1min"],
        ["耐电压 DWV", "150V, 1~5s", "耐压测试为额定电压 300%；50V×300%=150V，时间 1~5s，充放电电流 50mA max。", "Murata specifications: Test Voltage 300% of rated voltage; Applied Time 1s to 5s; current 50mA max。", "150V, 1~5s, 50mA max"],
        ["焊锡性", "245±5℃, 2±0.5s", "端头应 95% 均匀连续上锡。", "Murata solderability test: 245±5℃, 2±0.5s。", "Sn-3.0Ag-0.5Cu"],
    ]
    for row_idx, row in enumerate(rows, 3):
        for col_idx, value in enumerate(row, 1):
            cell = basis.cell(row_idx, col_idx, value)
            cell.font = Font(name="宋体", size=9, bold=(row_idx == 3))
            cell.alignment = Alignment(horizontal="center" if col_idx < 3 else "left", vertical="center", wrap_text=True)
            cell.border = border(thin)
            if row_idx == 3:
                cell.fill = PatternFill("solid", fgColor="EDEDED")

    wb.save(OUTPUT)


if __name__ == "__main__":
    build()
